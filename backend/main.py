from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from sqlalchemy.orm import Session
from datetime import datetime
import os
import re
import logging

logger = logging.getLogger("edutrack")
logging.basicConfig(level=logging.INFO)

# Import database and crud modules
from crud import (
    get_dashboard_stats, mark_attendance, get_attendance,
    add_marks, get_marks, get_at_risk_students,
    create_user, get_user_by_email, create_student, get_students,
    get_student_by_id, update_student, delete_student, get_student_stats,
    bulk_mark_attendance, process_bulk_attendance, process_bulk_attendance_pdf,
    process_bulk_marks, process_bulk_marks_excel,
    send_student_intervention_alert, get_subject_difficulty_analytics,
    predict_student_performance_ml, generate_student_gemini_insights
)
from schemas import (
    AttendanceCreate, MarksCreate, RiskPrediction,
    UserCreate, LoginSchema, StudentCreate, StudentUpdate
)
from database import SessionLocal, engine, Base
from auth import verify_password, create_access_token

# Import auth router file
import auth 

# 1. Initialize app
app = FastAPI(title="EduTrack AI")

# 2. Add CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "https://edutrack-ai-pink.vercel.app",
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 3. Include auth router
app.include_router(auth.router)

# 4. Create database tables
Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/")
def home():
    return {
        "message": "EduTrack AI Backend Running",
        "status": "online",
        "docs": "/docs"
    }

# ==========================================
# AUTHENTICATION
# ==========================================

@app.post("/register")
def register(user: UserCreate, db: Session = Depends(get_db)):
    existing = get_user_by_email(db, user.email)
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    return create_user(db, user)

@app.post("/login")
def login(user: LoginSchema, db: Session = Depends(get_db)):
    db_user = get_user_by_email(db, user.email)
    if not db_user or not verify_password(user.password, db_user.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_access_token({"sub": db_user.email})
    return {
        "access_token": token,
        "token_type": "bearer"
    }

@app.post("/token")
def token_login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    db_user = get_user_by_email(db, form_data.username)
    if not db_user or not verify_password(form_data.password, db_user.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_access_token({"sub": db_user.email})
    return {
        "access_token": token,
        "token_type": "bearer"
    }


# ==========================================
# STUDENTS DIRECTORY
# ==========================================

@app.get("/students")
def list_students(db: Session = Depends(get_db)):
    return get_students(db)

@app.post("/students")
def add_student(student: StudentCreate, db: Session = Depends(get_db)):
    return create_student(db, student)

@app.get("/students/stats")
def student_stats(db: Session = Depends(get_db)):
    return get_student_stats(db)

@app.get("/students/{student_id}")
def fetch_student_by_id(student_id: int, db: Session = Depends(get_db)):
    student = get_student_by_id(db, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@app.put("/students/{student_id}")
def edit_student(student_id: int, student_data: StudentUpdate, db: Session = Depends(get_db)):
    updated = update_student(db, student_id, student_data.dict(exclude_unset=True))
    if not updated:
        raise HTTPException(status_code=404, detail="Student not found")
    return updated

@app.delete("/students/{student_id}")
def delete_student_record(student_id: int, db: Session = Depends(get_db)):
    return delete_student(db, student_id)


# ==========================================
# ATTENDANCE MANAGEMENT
# ==========================================

@app.get("/attendance")
def attendance_list(db: Session = Depends(get_db)):
    return get_attendance(db)

@app.post("/attendance")
def add_attendance(attendance: AttendanceCreate, db: Session = Depends(get_db)):
    return mark_attendance(db, attendance)

@app.post("/attendance/bulk-mark")
def add_bulk_attendance(payload: dict, db: Session = Depends(get_db)):
    records = payload.get("records", [])
    return bulk_mark_attendance(db, records)

@app.post("/attendance/bulk-upload")
async def bulk_upload_attendance(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported.")
    
    contents = await file.read()
    decoded_contents = contents.decode('utf-8')
    return process_bulk_attendance(db, decoded_contents)

@app.post("/attendance/bulk-pdf-upload")
async def bulk_upload_attendance_pdf(file: UploadFile = File(...), subject: str = "General", db: Session = Depends(get_db)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported for this endpoint.")
    
    contents = await file.read()
    return process_bulk_attendance_pdf(db, contents, subject)


# ==========================================
# MARKS & EXAM MANAGEMENT
# ==========================================

@app.get("/marks")
def marks_list(db: Session = Depends(get_db)):
    return get_marks(db)

@app.post("/marks")
def create_marks(mark: MarksCreate, db: Session = Depends(get_db)):
    return add_marks(db, mark)

@app.post("/marks/bulk-upload")
async def bulk_upload_marks(
    file: UploadFile = File(...), 
    subject: str = None, 
    exam_type: str = None, 
    db: Session = Depends(get_db)
):
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported.")
    
    contents = await file.read()
    filename_clean = file.filename.replace('.xlsx', '').replace('.XLSX', '').replace('.csv', '').replace('.CSV', '')

    if not exam_type or exam_type.strip() in ["", "undefined", "null"]:
        exam_type = "T1"
        for part in re.split(r'[_ \-]', filename_clean.upper()):
            if part in ["T1", "T2", "T3", "T4", "MID", "END"]:
                exam_type = part
                break

    if not subject or subject.strip() in ["", "undefined", "null"]:
        parts = [p.strip() for p in re.split(r'[_ \-]', filename_clean) if p.strip()]
        ignore_words = {
            "T1", "T2", "T3", "T4", "MARKS", "MARKSHEET", "SCORES", "SY1", "SY2", "SY3", 
            "TY", "BTECH", "BATCH", "A", "B", "C", "D", "FINAL", "MID", "SEM", "SEM1", 
            "SEM2", "SEM3", "SEM4", "SEM5", "SEM6", "SEM7", "SEM8", "EXAM", "RESULT", "COMPILED"
        }
        subject_parts = [p for p in parts if p.upper() not in ignore_words and not p.isdigit()]
        
        if subject_parts:
            processed_parts = [p.upper() if len(p) <= 4 and p.isalpha() else p.capitalize() for p in subject_parts]
            subject = " ".join(processed_parts)
        else:
            subject = "General"

    if file.filename.endswith('.xlsx'):
        return process_bulk_marks_excel(db, contents, subject, exam_type)
    else:
        return process_bulk_marks(db, contents.decode('utf-8'), subject, exam_type)


# ==========================================
# DASHBOARD, ANALYTICS & REPORTS
# ==========================================

@app.get("/dashboard")
def dashboard(db: Session = Depends(get_db)):
    return get_dashboard_stats(db)

@app.get("/analytics/at-risk")
def at_risk_students(db: Session = Depends(get_db)):
    return get_at_risk_students(db)

@app.post("/predict-risk")
def predict_risk(data: RiskPrediction):
    if data.attendance < 60 or data.average_marks < 40:
        risk = "High"
        recommendation = "Immediate academic intervention required"
    elif data.attendance < 75 or data.average_marks < 60:
        risk = "Medium"
        recommendation = "Monitor student performance closely"
    else:
        risk = "Low"
        recommendation = "Student performing well"

    return {
        "risk_level": risk,
        "recommendation": recommendation
    }

@app.get("/analytics/ml-predict/{student_id}")
def get_ml_prediction(student_id: int, db: Session = Depends(get_db)):
    return predict_student_performance_ml(db, student_id)

@app.post("/analytics/send-alert/{student_id}")
def send_intervention_alert(student_id: int, db: Session = Depends(get_db)):
    return send_student_intervention_alert(db, student_id)

@app.get("/analytics/subject-heatmap")
def get_subject_heatmap(db: Session = Depends(get_db)):
    return get_subject_difficulty_analytics(db)

@app.get("/analytics/ai-insights/{student_id}")
def get_student_ai_insights(student_id: int, db: Session = Depends(get_db)):
    return generate_student_gemini_insights(db, student_id)

@app.get("/reports/compiled-marksheet")
@app.get("/reports/performance-excel")
def generate_compiled_marksheet(db: Session = Depends(get_db)):
    from models import Student, Marks
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compiled Marksheet"

    students = db.query(Student).all()
    all_marks = db.query(Marks).all()
    
    subjects = sorted(list(set([m.subject for m in all_marks if m.subject])))

    headers = ["Roll No / ID", "Student Name", "Email", "Attendance %"] + subjects + ["Total Score", "Percentage", "Status"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D3D3D3")

    row_num = 2
    for student in students:
        student_marks = [m for m in all_marks if m.student_id == student.id]
        if not student_marks:
            continue

        sheet.cell(row=row_num, column=1).value = student.id
        sheet.cell(row=row_num, column=2).value = student.name
        sheet.cell(row=row_num, column=3).value = student.email
        sheet.cell(row=row_num, column=4).value = student.attendance

        total_obtained = 0
        max_possible = len(subjects) * 100 
        marks_dict = {m.subject: m.marks for m in student_marks}
        
        col_idx = 5
        for sub in subjects:
            val = marks_dict.get(sub, "-")
            sheet.cell(row=row_num, column=col_idx).value = val
            if val != "-":
                total_obtained += val
            col_idx += 1
        
        percentage = (total_obtained / max_possible * 100) if max_possible > 0 else 0
        sheet.cell(row=row_num, column=col_idx).value = total_obtained
        sheet.cell(row=row_num, column=col_idx+1).value = round(percentage, 2)
        sheet.cell(row=row_num, column=col_idx+2).value = "Pass" if percentage >= 40 else "Fail"
        
        status_cell = sheet.cell(row=row_num, column=col_idx+2)
        if percentage >= 40:
            status_cell.fill = PatternFill(fill_type="solid", fgColor="90EE90")
        else:
            status_cell.fill = PatternFill(fill_type="solid", fgColor="FF9999")

        row_num += 1

    for col in ["B", "C"]:
        sheet.column_dimensions[col].width = 25

    filename = f"Compiled_Marksheet_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    workbook.save(filename)
    
    return FileResponse(
        path=filename, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        filename=filename
    )


# ==========================================
# CALENDAR & TIMETABLE PDF HELPERS
# ==========================================

@app.post("/calendar/upload-pdf")
async def upload_calendar_pdf(file: UploadFile = File(...)):
    return {"message": f"Successfully received calendar PDF '{file.filename}'"}

@app.post("/timetable/upload-pdf")
async def upload_timetable_pdf(file: UploadFile = File(...)):
    return {"message": f"Successfully received timetable PDF '{file.filename}'"}